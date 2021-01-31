import subprocess
import sys

# This uses system terminal to install python modules
subprocess.call([sys.executable, "-m", "pip", "install", 'XlsxWriter'])
subprocess.call([sys.executable, "-m", "pip", "install", 'py3-validate-email'])
subprocess.call([sys.executable, "-m", "pip", "install", 'openpyxl'])

import os
import xlsxwriter as xwrite
import openpyxl as pyxl
from Classes import Person, Department


def init():
    # opening database file
    database_loc = os.path.join(sys.path[0], 'database.xlsx')
    try:
        database_wb = pyxl.load_workbook(database_loc)
        sheets = database_wb
    except:
        print("\n" + "Error: No Excel Sheet Named 'database.xlsx' Found :( ")
        input("Please paste the database sheet into this folder, and run the script again!")

    # creating diagnostic sheets
    diagnostic_loc = os.path.join(sys.path[0], 'Actions Needed.xlsx')
    diagnostic_wb = xwrite.Workbook(diagnostic_loc)
    website_sheet = diagnostic_wb.add_worksheet('Update Invalid Websites')
    email_sheet = diagnostic_wb.add_worksheet('Fix Broken Emails')
    status_sheet = diagnostic_wb.add_worksheet('Remove Inactive Faculty')
    date_sheet = diagnostic_wb.add_worksheet('Renew Outdated Entries')
    duplicates_sheet = diagnostic_wb.add_worksheet('Revise Duplicates')
    missing_info_sheet = diagnostic_wb.add_worksheet('Fill-in Missing Information')

    # sheet formats
    bold = diagnostic_wb.add_format({'bold': True, 'font_color': 'red'})

    # turning database into data
    departments, num_faculty = make_person(sheets)

    # PERFORMING DIAGNOSTICS
    # ///////////////////////////////////////////////////////////////
    # checking websites
    website_tasks = check_websites(departments, website_sheet, bold)
    # checking email
    email_tasks = check_email(departments, email_sheet, bold)
    # checking status
    status_tasks = check_status(departments, status_sheet, bold)
    # checking date modified
    date_tasks = check_date(departments, date_sheet, bold)
    # checking duplicates
    duplicates_tasks = check_duplicates(departments, duplicates_sheet, bold)
    # checking missing information
    missing_info_tasks = check_missing_info(departments, missing_info_sheet, bold)
    # ///////////////////////////////////////////////////////////////
    diagnostic_wb.close()

    # calculating percentage of update
    total_tasks = website_tasks + email_tasks + status_tasks + date_tasks + duplicates_tasks + missing_info_tasks
    percentage = 100 * (1 - round(total_tasks / num_faculty, 4))

    # printing diagnostic result
    print("\n" + "ACTIONS NEEDED:" + "\n" + "" + "\n"
          + "Website Tasks = " + str(website_tasks) + "\n"
          + "Email Tasks = " + str(email_tasks) + "\n"
          + "Status Tasks = " + str(status_tasks) + "\n"
          + "Date Tasks = " + str(date_tasks) + "\n"
          + "Duplicates Tasks = " + str(duplicates_tasks) + "\n"
          + "Missing Info Tasks = " + str(missing_info_tasks) + "\n" + "" + "\n"
          + "(Please see the 'Actions Needed.xlsx' sheet for details)" + "\n" + "" + "\n"
          + "Database " + str(percentage) + "% Updated." + "\n")

    # exit
    input("\n" + "Completed! Press any key to exit... ")
    exit()


def make_person(sheets):
    """Returns a list and an integer, where
    list: Department objects, which each has a list of Person objects;
    integer: the total number of faculty members"""

    # status
    print("\n" + "Making People...")

    # list that stores people
    departments = []
    num_faculty = 0

    for sheet in sheets:

        people = []
        # openpyxl uses excel number conventions (starts from 1)
        for i in range(2, get_max_row(sheet)):
            print(get_max_row(sheet))
            if sheet.cell(i, 1) != "":
                last_name = sheet.cell(i, 1).value
                first_name = sheet.cell(i, 2).value
                institution = sheet.cell(i, 3).value
                program = sheet.cell(i, 4).value
                position = sheet.cell(i, 5).value
                knowledge = sheet.cell(i, 6).value
                email = sheet.cell(i, 7).value
                website = sheet.cell(i, 8).value
                gender = sheet.cell(i, 9).value
                urm = sheet.cell(i, 10).value
                date_modified = sheet.cell(i, 11).value
                status = sheet.cell(i, 12).value

                new_person = Person(last_name, first_name, institution, program,
                                    position, knowledge, email, website, gender, urm, date_modified, status)
                people.append(new_person)

                # status
                print(new_person.name + "|| completed")

        new_department = Department(sheet.title, people)
        departments.append(new_department)
        num_faculty += len(people)

        # status
        print(new_department.name + "|| completed")

    return departments, num_faculty


from urllib.request import urlopen
from urllib.error import HTTPError, URLError


def check_websites(departments, website_sheet, style, any_false=[], i=0):
    """Returns an integer: the number of inaccessible websites"""

    # status
    print("\n" + "Checking Websites..." + "\n" + "" + "\n"
          + "This will take a while, " + "\n"
          + "Please do not press any key during this process!" + "\n")

    # check if a website is valid
    for department in departments:
        website_sheet.write(i, 0, department.name, style)
        i += 1
        for person in department.people:
            try:
                response = urlopen(person.website)
                if response.code != 200:
                    print("Webpage Error: " + person.name)
                    any_false.append(False)
            except HTTPError as e:
                codes = [403, 999]
                if e.code not in codes:
                    print("Webpage Error: " + person.name + " || " + str(e.code) + "http")
                    website_sheet.write(i, 0, person.name)
                    i += 1
                    any_false.append(False)
            except URLError as e:
                if "[SSL: CERTIFICATE_VERIFY_FAILED]" not in str(e.args):
                    print("Webpage Error: " + person.name + " || " + str(e.args) + "url")
                    website_sheet.write(i, 0, person.name)
                    i += 1
                    any_false.append(False)
            except:
                pass

    if all(any_false):
        print("\n" + "Congrats, All websites are working! ")
        return 0

    return i - len(departments)


from validate_email import validate_email


def check_email(departments, email_sheet, style, any_false=[], i=0):
    """Returns an integer: the number of invalid email addresses"""

    # status
    print("\n" + "Checking emails...")

    # check if a website is valid
    for department in departments:
        email_sheet.write(i, 0, department.name, style)
        i += 1
        for person in department.people:
            if person.email != "n/a":
                try:
                    is_valid = validate_email(person.email, check_regex=True, check_mx=False)
                    if is_valid is False:
                        print("Email Error: " + person.name)
                        email_sheet.write(i, 0, person.name)
                        i += 1
                        any_false.append(is_valid)
                except:
                    pass

    if all(any_false):
        print("\n" + "Congrats, All emails are valid! ")
        return 0

    return i - len(departments)


def check_status(departments, status_sheet, style, any_false=[], i=0):
    """Returns an integer: the number of inactive faculty members"""

    # status
    print("\n" + "Checking status...")

    for department in departments:
        status_sheet.write(i, 0, department.name, style)
        i += 1
        for person in department.people:
            non_active = ['Inactive', 'Unsure', 'Transferred']
            if person.status in non_active:
                print("Inactive Status: " + person.name)
                status_sheet.write(i, 0, person.name)
                i += 1
                any_false.append(False)

    if all(any_false):
        print("\n" + "Congrats, All faculty members are active! ")
        return 0

    return i - len(departments)


from datetime import datetime


def check_date(departments, date_sheet, style, any_false=[], i=0, year=365):
    """Returns an integer: the number of outdated faculty members"""

    # status
    print("\n" + "Checking dates...")

    for department in departments:
        date_sheet.write(i, 0, department.name, style)
        i += 1
        for person in department.people:

            if person.date_modified == "":
                print("Missing Date: " + person.name)
                date_sheet.write(i, 0, person.name)
                i += 1
                any_false.append(False)

            else:
                try:
                    calculate = lambda d1, d2: abs((d1 - d2).days)
                    today = datetime.today()
                    last_date = person.date_modified
                    period = calculate(today, last_date)

                    if period > year:
                        print("Outdated Entry: " + person.name)
                        date_sheet.write(i, 0, person.name)
                        i += 1
                        any_false.append(False)
                except:
                    print("Invalid Entry: " + person.name)
                    date_sheet.write(i, 0, person.name)
                    i += 1
                    any_false.append(False)

    if all(any_false):
        print("\n" + "Congrats, All people are up-to-date! ")
        return 0

    return i - len(departments)


def check_duplicates(departments, duplicates_sheet, style, any_false=[], i=0):
    """Returns an integer: the number of duplicates within departments"""

    # status
    print("\n" + "Checking duplicates...")

    for department in departments:
        duplicates_sheet.write(i, 0, department.name, style)
        i += 1

        unique_persons = {}
        for person in department.people:
            if person.name not in unique_persons:
                unique_persons[person.name] = 1
            else:
                print("Duplicated Entry: " + person.name)
                duplicates_sheet.write(i, 0, person.name)
                i += 1
                any_false.append(False)

    if all(any_false):
        print("\n" + "Congrats, All people are up-to-date! ")
        return 0

    return i - len(departments)


def check_missing_info(departments, missing_info_sheet, style, any_false=[], i=0):
    """Returns an integer: the number of faculty member with missing information"""

    # status
    print("\n" + "Checking missing information...")

    for department in departments:
        missing_info_sheet.write(i, 0, department.name, style)
        i += 1

        for person in department.people:
            has_missing_info = False
            for attribute in person.all_attributes:
                if attribute is None:
                    has_missing_info = True
                    break
            if has_missing_info:
                missing_info_sheet.write(i, 0, person.name)
                print("Missing Information: " + person.name)
                any_false.append(False)
                i += 1

    if all(any_false):
        print("\n" + "Congrats, All people have information! ")
        return 0

    return i - len(departments)


# Helper Functions


def get_max_row(sheet):
    """Returns an integer: the number of rows in sheet"""

    empty_rows = 0
    max_row = 0

    while empty_rows < 10:
        max_row += 1
        if sheet.cell(max_row, 1).value == "" or sheet.cell(max_row, 1).value is None:
            empty_rows += 1
        else:
            empty_rows = 0

    max_row -= 10

    return max_row


init()
